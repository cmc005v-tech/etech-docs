# -*- coding: utf-8 -*-
"""
生成学习网站 public/resources 下的 8 个真实资源文件：
  - 5 份第 1 课教案 PDF（源：跨境电商课程开发/各课逐课教案文稿，经 HTML 由 Edge 无头打印）
  - 2 个 XLSX 模板（物流成本拆解表 / 供应商评估打分表，源：供应链课程配套工具包）
  - 1 个工具包样例 ZIP（聚合多课程配套工具样例）
用法：python scripts/generate_resource_files.py
"""
import csv
import html
import io
import re
import subprocess
import sys
import zipfile
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

ROOT = Path(__file__).resolve().parents[1]
SRC = ROOT.parent / "跨境电商课程开发"
OUT = ROOT / "docs" / "public" / "resources"
TMP = ROOT / ".gen_tmp"
EDGE = r"C:\Program Files (x86)\Microsoft\Edge\Application\msedge.exe"

# ---------------------------------------------------------------- md → html

def inline(text: str) -> str:
    text = html.escape(text)
    text = re.sub(r"\*\*(.+?)\*\*", r"<strong>\1</strong>", text)
    text = re.sub(r"`(.+?)`", r"<code>\1</code>", text)
    return text


def md_to_html(md: str) -> str:
    """轻量 Markdown 转换器，覆盖教案文稿用到的语法。"""
    out, lines = [], md.splitlines()
    i, n = 0, len(lines)
    while i < n:
        line = lines[i]
        s = line.strip()
        # 代码块
        if s.startswith("```"):
            buf = []
            i += 1
            while i < n and not lines[i].strip().startswith("```"):
                buf.append(lines[i])
                i += 1
            out.append("<pre>" + html.escape("\n".join(buf)) + "</pre>")
            i += 1
            continue
        # 表格
        if s.startswith("|"):
            tbl = []
            while i < n and lines[i].strip().startswith("|"):
                cells = [c.strip() for c in lines[i].strip().strip("|").split("|")]
                if not all(re.fullmatch(r":?-{3,}:?", c) for c in cells):
                    tbl.append(cells)
                i += 1
            if tbl:
                rows = "".join(
                    "<tr>" + "".join(f"<td>{inline(c)}</td>" for c in r) + "</tr>"
                    for r in tbl
                )
                out.append(f"<table>{rows}</table>")
            continue
        # 标题
        m = re.match(r"^(#{1,6})\s+(.*)", s)
        if m:
            lv = len(m.group(1))
            out.append(f"<h{lv}>{inline(m.group(2))}</h{lv}>")
            i += 1
            continue
        # 引用
        if s.startswith(">"):
            buf = []
            while i < n and lines[i].strip().startswith(">"):
                buf.append(lines[i].strip().lstrip("> ").rstrip())
                i += 1
            out.append("<blockquote>" + "<br>".join(inline(b) for b in buf if b) + "</blockquote>")
            continue
        # 无序列表
        if re.match(r"^[-*]\s+", s):
            items = []
            while i < n and re.match(r"^[-*]\s+", lines[i].strip()):
                items.append(re.sub(r"^[-*]\s+", "", lines[i].strip()))
                i += 1
            out.append("<ul>" + "".join(f"<li>{inline(x)}</li>" for x in items) + "</ul>")
            continue
        # 有序列表
        if re.match(r"^\d+[.、]\s*", s):
            items = []
            while i < n and re.match(r"^\d+[.、]\s*", lines[i].strip()):
                items.append(re.sub(r"^\d+[.、]\s*", "", lines[i].strip()))
                i += 1
            out.append("<ol>" + "".join(f"<li>{inline(x)}</li>" for x in items) + "</ol>")
            continue
        # 分隔线 / 空行
        if s in ("---", "***"):
            out.append("<hr>")
            i += 1
            continue
        if not s:
            i += 1
            continue
        out.append(f"<p>{inline(s)}</p>")
        i += 1
    return "\n".join(out)


CSS = """
@page { size: A4; margin: 18mm 16mm; }
* { box-sizing: border-box; }
body { font-family: "Microsoft YaHei","微软雅黑","PingFang SC",sans-serif;
       font-size: 10.5pt; line-height: 1.75; color: #1f2937; margin: 0; }
.cover { border: 2px solid #2b6cb0; border-radius: 8px; padding: 26px 30px; margin-bottom: 22px;
         background: #f0f6fc; page-break-after: avoid; }
.cover .tag { display: inline-block; background: #2b6cb0; color: #fff; font-size: 9pt;
              padding: 2px 10px; border-radius: 999px; margin-bottom: 10px; }
.cover h1 { font-size: 17pt; margin: 6px 0; color: #1a365d; }
.cover .meta { color: #4a5568; font-size: 9.5pt; }
h1 { font-size: 15pt; color: #1a365d; border-bottom: 2px solid #2b6cb0; padding-bottom: 6px; }
h2 { font-size: 13pt; color: #2b6cb0; margin-top: 22px; }
h3 { font-size: 11.5pt; color: #2c5282; margin-top: 18px; }
h4 { font-size: 10.5pt; }
table { border-collapse: collapse; width: 100%; margin: 10px 0; font-size: 9.5pt; }
td, th { border: 1px solid #cbd5e0; padding: 5px 8px; text-align: left; }
tr:nth-child(even) td { background: #f7fafc; }
blockquote { border-left: 3px solid #63b3ed; background: #ebf8ff; margin: 10px 0;
             padding: 8px 14px; color: #2c5282; }
pre { background: #f1f5f9; border: 1px solid #e2e8f0; border-radius: 6px;
      padding: 10px 14px; font-family: Consolas,"Microsoft YaHei",monospace;
      font-size: 9pt; white-space: pre-wrap; }
ul, ol { padding-left: 22px; margin: 8px 0; }
li { margin: 3px 0; }
hr { border: none; border-top: 1px dashed #cbd5e0; margin: 18px 0; }
code { background: #edf2f7; padding: 1px 5px; border-radius: 3px; font-size: 9pt; }
.footer { margin-top: 28px; padding-top: 10px; border-top: 1px solid #cbd5e0;
          font-size: 8.5pt; color: #718096; }
"""


def build_pdf(md_src: Path, course_name: str, course_code: str, out_pdf: Path):
    text = md_src.read_text(encoding="utf-8")
    m = re.search(r"(# 第1课教案：.*?)(?=\n# 第2课教案|\Z)", text, re.S)
    if not m:
        raise RuntimeError(f"未找到第1课教案: {md_src.name}")
    lesson_md = m.group(1)
    lesson_title = lesson_md.splitlines()[0].lstrip("# ").strip()
    body = md_to_html(lesson_md.splitlines()[0] and "\n".join(lesson_md.splitlines()[1:]))
    page = f"""<!DOCTYPE html><html lang="zh-CN"><head><meta charset="utf-8">
<style>{CSS}</style></head><body>
<div class="cover">
  <span class="tag">跨境电商高级实战系列 · 试听样章</span>
  <h1>{html.escape(course_name)}</h1>
  <div class="meta">课程代码：{course_code} ｜ 本文档为第 1 课完整教案（讲师版样例）</div>
</div>
{body}
<div class="footer">跨境电商高级实战系列 · 学习网站 ｜ 教案样例仅供试听评估使用</div>
</body></html>"""
    tmp_html = TMP / (out_pdf.stem + ".html")
    tmp_html.write_text(page, encoding="utf-8")
    out_pdf.parent.mkdir(parents=True, exist_ok=True)
    subprocess.run(
        [EDGE, "--headless=new", "--disable-gpu", "--no-pdf-header-footer",
         f"--print-to-pdf={out_pdf}", tmp_html.as_uri()],
        check=True, capture_output=True, timeout=120,
    )
    print(f"[PDF] {out_pdf.name}  {out_pdf.stat().st_size/1024:.0f} KB")

# ---------------------------------------------------------------- XLSX 样式

THIN = Border(*[Side(style="thin", color="B0BEC5")] * 4)
HDR_FILL = PatternFill("solid", fgColor="2B6CB0")
GRP_FILL = PatternFill("solid", fgColor="EBF4FB")
HDR_FONT = Font(name="微软雅黑", size=10, bold=True, color="FFFFFF")
GRP_FONT = Font(name="微软雅黑", size=10, bold=True, color="1A365D")
BODY_FONT = Font(name="微软雅黑", size=10)
CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
LEFT = Alignment(horizontal="left", vertical="center", wrap_text=True)


def style_row(ws, row, ncols, font, fill=None, align=LEFT):
    for c in range(1, ncols + 1):
        cell = ws.cell(row=row, column=c)
        cell.font = font
        cell.border = THIN
        cell.alignment = align
        if fill:
            cell.fill = fill


def title_block(ws, title, subtitle, ncols):
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=ncols)
    ws.cell(row=1, column=1, value=title).font = Font(name="微软雅黑", size=14, bold=True, color="1A365D")
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=ncols)
    ws.cell(row=2, column=1, value=subtitle).font = Font(name="微软雅黑", size=9, color="718096")


def build_logistics_cost_xlsx(path: Path):
    """全链路成本拆解表（工具一）：直邮 vs 海外仓对比，含自动汇总公式。"""
    wb = Workbook()
    ws = wb.active
    ws.title = "全链路成本拆解"
    title_block(ws, "跨境电商全链路成本拆解表",
                "供应链管理课程配套工具（工具一）· 填写单件成本后自动汇总，对比直邮与海外仓模式", 4)
    headers = ["成本项目", "直邮模式（元/件）", "海外仓模式（元/件）", "备注"]
    rows = [
        ("一、采购 / 生产成本", None, None, None, True),
        ("工厂出厂价", None, None, "含包装"),
        ("质检费用（分摊）", None, None, ""),
        ("二、头程物流成本", None, None, None, True),
        ("国内集货运输", None, None, "工厂→集货仓"),
        ("头程国际运输", None, None, "空运/海运/快递"),
        ("保险费用", None, None, ""),
        ("三、关税与合规成本", None, None, None, True),
        ("进口关税", None, None, "按目标市场税率"),
        ("VAT / 销售税", None, None, "可抵扣部分标注"),
        ("清关服务费", None, None, ""),
        ("产品认证分摊", None, None, "CE/FCC/FDA 等"),
        ("四、仓储成本", None, None, None, True),
        ("国内仓租（分摊）", None, None, ""),
        ("海外仓月租（分摊）", None, None, "按件分摊"),
        ("FBA 费用（如适用）", None, None, ""),
        ("五、尾程配送成本", None, None, None, True),
        ("本地快递费", None, None, ""),
        ("包装耗材", None, None, ""),
        ("六、逆向物流成本", None, None, None, True),
        ("退货处理费（分摊）", None, None, ""),
        ("弃货损失（分摊）", None, None, ""),
        ("翻新 / 重新上架费", None, None, ""),
    ]
    ws.append([]); ws.append(headers)
    hr = 4  # headers row
    style_row(ws, hr, 4, HDR_FONT, HDR_FILL, CENTER)
    data_rows = []
    for item in rows:
        group = len(item) > 4 and item[4]
        r = ws.max_row + 1
        ws.cell(row=r, column=1, value=item[0])
        if group:
            style_row(ws, r, 4, GRP_FONT, GRP_FILL)
        else:
            ws.cell(row=r, column=4, value=item[3])
            style_row(ws, r, 4, BODY_FONT)
            data_rows.append(r)
    # 汇总
    first, last = data_rows[0], data_rows[-1]
    total_r = ws.max_row + 1
    ws.cell(row=total_r, column=1, value="单件总成本")
    for col, letter in ((2, "B"), (3, "C")):
        parts = "+".join(f"{letter}{r}" for r in data_rows)
        ws.cell(row=total_r, column=col, value=f"=IF(COUNT({letter}{first}:{letter}{last})=0,\"\",{parts})")
    style_row(ws, total_r, 4, Font(name="微软雅黑", size=10, bold=True), PatternFill("solid", fgColor="FFF5CC"))
    ratio_r = total_r + 1
    ws.cell(row=ratio_r, column=1, value="售价（元/件）")
    ws.cell(row=ratio_r + 1, column=1, value="成本占售价比例")
    for col, letter in ((2, "B"), (3, "C")):
        ws.cell(row=ratio_r + 1, column=col,
                value=f'=IF(OR({letter}{ratio_r}="",{letter}{ratio_r}=0),"",{letter}{total_r}/{letter}{ratio_r})')
        ws.cell(row=ratio_r + 1, column=col).number_format = "0.0%"
    for r in (ratio_r, ratio_r + 1):
        style_row(ws, r, 4, BODY_FONT)
    note_r = ratio_r + 3
    ws.cell(row=note_r, column=1, value="关键对比指标：① 直邮 vs 海外仓成本差额　② 关税成本占比变化　③ 综合成本优化空间（目标：降低 30%-40%）").font = Font(name="微软雅黑", size=9, color="718096")
    for col, w in ((1, 26), (2, 18), (3, 18), (4, 22)):
        ws.column_dimensions[get_column_letter(col)].width = w

    ws2 = wb.create_sheet("头程方式参考基准")
    title_block(ws2, "头程运输方式成本与时效基准（参考值）", "数据口径见课程第 1 课全链路结构解析，实际报价以货代为准", 5)
    ws2.append([]); ws2.append(["运输方式", "成本（元/kg）", "时效", "适用场景", "风险提示"])
    style_row(ws2, 4, 5, HDR_FONT, HDR_FILL, CENTER)
    for row in [
        ("空运", "30-50", "5-7 天", "高货值 / 急补货 / 新品测款", "旺季舱位紧张，价格波动大"),
        ("海运", "5-8", "25-35 天", "大批量 / 常规补货", "需预留长备货周期，压资金"),
        ("国际快递", "50-80", "3-5 天", "样品 / 小件急件", "成本最高，仅应急使用"),
        ("中欧铁路", "12-20", "18-25 天", "欧洲市场 / 性价比折中", "口岸拥堵影响时效"),
    ]:
        ws2.append(row)
        style_row(ws2, ws2.max_row, 5, BODY_FONT)
    for col, w in ((1, 14), (2, 16), (3, 12), (4, 30), (5, 30)):
        ws2.column_dimensions[get_column_letter(col)].width = w
    wb.save(path)
    print(f"[XLSX] {path.name}  {path.stat().st_size/1024:.0f} KB")


def build_supplier_xlsx(path: Path):
    """供应商评估打分表（工具四）：四维加权评分 + 自动分级。"""
    wb = Workbook()
    ws = wb.active
    ws.title = "评估打分表"
    title_block(ws, "供应商评估打分表",
                "供应链管理课程配套工具（工具四）· 四维加权评分：交付质量40% + 交期稳定性30% + 成本竞争力20% + 配合度10%", 6)
    ws.append([]); ws.append(["供应商名称", "", "评估日期", "", "评估人", ""])
    style_row(ws, 4, 6, BODY_FONT)
    for c in (2, 4, 6):
        ws.cell(row=4, column=c).fill = PatternFill("solid", fgColor="FFFBEA")
    ws.append([])
    ws.append(["评估维度", "权重", "评分（满分100）", "加权得分", "评分要点", "备注"])
    style_row(ws, 6, 6, HDR_FONT, HDR_FILL, CENTER)
    dims = [
        ("交付质量", 0.4, "出货合格率、退货率、品质稳定性"),
        ("交期稳定性", 0.3, "准时交付率、延期天数、旺季保障"),
        ("成本竞争力", 0.2, "价格水平、年度降价意愿、付款条件"),
        ("配合度", 0.1, "响应速度、验厂配合、问题解决能力"),
    ]
    start = 7
    for idx, (name, w, key) in enumerate(dims):
        r = start + idx
        ws.cell(row=r, column=1, value=name)
        ws.cell(row=r, column=2, value=w).number_format = "0%"
        ws.cell(row=r, column=4, value=f"=IF(C{r}=\"\",\"\",C{r}*B{r})")
        ws.cell(row=r, column=5, value=key)
        style_row(ws, r, 6, BODY_FONT)
        ws.cell(row=r, column=2).alignment = CENTER
        ws.cell(row=r, column=4).alignment = CENTER
    total_r = start + len(dims)
    ws.cell(row=total_r, column=1, value="综合得分")
    ws.cell(row=total_r, column=4, value=f"=IF(COUNT(C{start}:C{total_r-1})=0,\"\",SUM(D{start}:D{total_r-1}))")
    ws.cell(row=total_r + 1, column=1, value="供应商等级")
    ws.cell(row=total_r + 1, column=4,
            value=f'=IF(D{total_r}="","",IF(D{total_r}>=85,"A级·战略供应商",IF(D{total_r}>=70,"B级·常规供应商",IF(D{total_r}>=60,"C级·观察供应商","D级·淘汰供应商"))))')
    for r in (total_r, total_r + 1):
        style_row(ws, r, 6, Font(name="微软雅黑", size=10, bold=True), PatternFill("solid", fgColor="FFF5CC"))
    ws.cell(row=total_r + 3, column=1,
            value="管理规则：A级≥85分 优先分配订单、季度review ｜ B级70-84分 正常合作、月度跟踪 ｜ C级60-69分 减少订单、30天改善 ｜ D级<60分 停止新订单、60天完成转移。核心品类必须保持≥2家合格供应商（主供70% / 备供30%）。"
            ).font = Font(name="微软雅黑", size=9, color="718096")
    for col, w in ((1, 16), (2, 10), (3, 16), (4, 14), (5, 40), (6, 16)):
        ws.column_dimensions[get_column_letter(col)].width = w

    ws2 = wb.create_sheet("评分细则")
    title_block(ws2, "评分细则（各维度满分100，按实际表现落分）", "来源：供应链管理课程配套工具包 · 工具四", 2)
    ws2.append([]); ws2.append(["维度", "评分标准"])
    style_row(ws2, 4, 2, HDR_FONT, HDR_FILL, CENTER)
    rules = [
        ("交付质量（40分）", "出货合格率≥99.5%：36-40分 ｜ 98%-99.5%：28-35分 ｜ 95%-98%：20-27分 ｜ <95%：0-19分"),
        ("交期稳定性（30分）", "准时交付率≥98%：27-30分 ｜ 90%-98%：21-26分 ｜ 80%-90%：15-20分 ｜ <80%：0-14分"),
        ("成本竞争力（20分）", "低于市场均价5%以上+月结：18-20分 ｜ 接近均价+月结：14-17分 ｜ 高于均价但品质好：10-13分 ｜ 价高且付款条件差：0-9分"),
        ("配合度（10分）", "24小时内响应+主动解决问题：9-10分 ｜ 48小时内响应：7-8分 ｜ 响应慢但能完成：5-6分 ｜ 配合度差：0-4分"),
    ]
    for name, rule in rules:
        ws2.append([name, rule])
        style_row(ws2, ws2.max_row, 2, BODY_FONT)
    ws2.column_dimensions["A"].width = 22
    ws2.column_dimensions["B"].width = 95
    wb.save(path)
    print(f"[XLSX] {path.name}  {path.stat().st_size/1024:.0f} KB")

# ---------------------------------------------------------------- ZIP 工具包

def build_toolkit_zip(path: Path):
    tmp = TMP / "toolkit"
    if tmp.exists():
        for f in tmp.rglob("*"):
            if f.is_file():
                f.unlink()
    tmp.mkdir(parents=True, exist_ok=True)

    # 1) 说明文件
    readme = """# 跨境电商高级实战系列 · 配套工具包样例

本压缩包为课程配套工具模板的**样例合集**，包含 5 件核心工具模板。

## 文件清单

| 文件 | 对应工具 | 来源课程 |
|------|---------|---------|
| 01-全链路成本拆解表.xlsx | 工具一：全链路成本拆解表 | 供应链管理（SCM-101） |
| 02-SKU分级管理SOP.md | 工具二：SKU 分级管理 SOP 模板 | 供应链管理（SCM-101） |
| 03-补货参数记录表.csv | 工具三：补货计划计算模型 | 供应链管理（SCM-101） |
| 04-供应商评估打分表.xlsx | 工具四：供应商评估打分表 | 供应链管理（SCM-101） |
| 05-风险预警阈值.csv | 工具五：供应链风险预警看板 | 供应链管理（SCM-101） |

## 使用建议
- Excel 模板含自动汇总 / 加权评分公式，填入数据即得结果；
- 参数与阈值（安全库存天数、分级分数线等）请结合自身企业体量调整；
- 各门知识库课程均配有独立工具包（6-10 件），报名后随课程资料一并分发。

跨境电商高级实战系列 · 学习网站
"""
    (tmp / "00-工具包说明.md").write_text(readme, encoding="utf-8")

    # 2/4) 复用两个 XLSX
    build_logistics_cost_xlsx(tmp / "01-全链路成本拆解表.xlsx")
    build_supplier_xlsx(tmp / "04-供应商评估打分表.xlsx")

    # 3) SKU 分级 SOP
    sku_sop = """# SKU 分级管理 SOP 模板（工具二 · 样例）

## 一、分级标准

| 等级 | 动销率标准 | 营收贡献 | 管理策略 |
|------|----------|---------|---------|
| A级 | ≥60% | 核心贡献（前70%） | 海外仓主力 + 安全库存 15 天 |
| B级 | 25%-60% | 重要补充 | 小批量海外试备 + 直发补充 |
| C级 | 5%-25% | 长尾 | 仅直发或评估清退 |
| D级 | <5% | 几乎无贡献 | 立即清退 |

## 二、月度执行流程
1. 每月 1 日：导出全量 SKU 近 30 天动销数据
2. 按动销率自动分级，生成升降级建议
3. 每月 3 日前：供应链主管审核
4. 执行调整：A→B 降库存天数；B→C 停海外仓补货转直发；C→D 启动清退；升级反向操作

## 三、升降级触发条件

| 变动 | 触发条件 | 执行动作 |
|------|---------|---------|
| C→B | 连续 14 天日均销量 ≥ B 级下限 | 恢复海外仓备货 |
| B→A | 连续 30 天日均销量 ≥ A 级下限且趋势稳定 | 增加安全库存 |
| 任意→D | 45 天无动销 | 立即启动清退 |

## 四、清退执行流程（D 类 SKU，目标 30 天消化 90%+）
- 第 1-7 天：站内捆绑促销，目标消化 30%
- 第 8-14 天：站外折扣渠道（Outlet/批发），目标消化 40%
- 第 15-21 天：尾货批发处理，目标消化 20%
- 第 22-30 天：剩余弃货或捐赠，完成清退
"""
    (tmp / "02-SKU分级管理SOP.md").write_text(sku_sop, encoding="utf-8")

    # 5) 补货参数记录表（CSV 样例，含公式说明行）
    buf = io.StringIO()
    w = csv.writer(buf)
    w.writerow(["# 补货计划计算模型（工具三·样例）核心公式："])
    w.writerow(["# 预测日均销量 = 近30天日均×0.6 + 近60天日均×0.3 + 季节性调整×0.1"])
    w.writerow(["# 安全库存量 = 预测日均销量×安全库存天数（A级15天 / B级10天）"])
    w.writerow(["# 补货触发点 = 安全库存量 + 预测日均销量×头程运输周期天数"])
    w.writerow(["# 补货目标量 = 预测日均销量×(头程周期+补货覆盖天数) - 当前库存 - 在途库存"])
    w.writerow(["SKU编码", "等级", "近30天日均", "近60天日均", "季节系数",
                "预测日均", "安全库存", "触发点", "当前库存", "在途库存", "是否触发补货"])
    w.writerow(["SKU-DEMO-001", "A", 20, 18, 1.1, "", "", "", "", "", ""])
    (tmp / "03-补货参数记录表.csv").write_text(buf.getvalue(), encoding="utf-8-sig")

    # 6) 风险预警阈值
    buf = io.StringIO()
    w = csv.writer(buf)
    w.writerow(["维度", "指标", "绿色（正常）", "黄色（预警）", "红色（警报）"])
    for row in [
        ("政策", "关税税率变化", "维持现状", "有征求意见稿", "正式生效"),
        ("政策", "平台规则更新", "月度≤1项", "月度2-3项", "月度>3项"),
        ("物流", "海运准班率", "≥85%", "70%-85%", "<70%"),
        ("物流", "空运价格周涨幅", "<5%", "5%-10%", ">10%"),
        ("物流", "尾程时效达标率", "≥97%", "95%-97%", "<95%"),
        ("库存", "库存周转天数", "≤60天", "60-90天", ">90天"),
        ("库存", "库龄>180天SKU占比", "≤10%", "10%-15%", ">15%"),
        ("库存", "断货率", "≤10%", "10%-15%", ">15%"),
        ("财务", "现金流周期", "≤90天", "90-120天", ">120天"),
        ("财务", "物流成本占比季度变化", "<1个百分点", "1-2个百分点", ">2个百分点"),
        ("竞争", "BSR排名月度变化", "稳定或上升", "下降10%-20%", "下降>20%"),
        ("竞争", "自然流量月度变化", "稳定或上升", "下降15%-30%", "下降>30%"),
    ]:
        w.writerow(row)
    w.writerow([])
    w.writerow(["响应机制", "绿色：月度常规监控", "黄色：周度跟踪，2周内启动改善计划", "红色：每日跟踪，48小时制定方案、72小时启动执行", ""])
    (tmp / "05-风险预警阈值.csv").write_text(buf.getvalue(), encoding="utf-8-sig")

    path.parent.mkdir(parents=True, exist_ok=True)
    with zipfile.ZipFile(path, "w", zipfile.ZIP_DEFLATED) as z:
        for f in sorted(tmp.iterdir()):
            z.write(f, f.name)
    print(f"[ZIP] {path.name}  {path.stat().st_size/1024:.0f} KB")


def main():
    TMP.mkdir(exist_ok=True)
    lessons = [
        ("跨境电商供应链课程-逐课教案文稿（第1-4课）.md", "跨境电商供应链管理", "SCM-101", "supply-chain"),
        ("跨境电商平台运营课程-逐课教案文稿（第1-4课）.md", "跨境电商平台运营", "PO-101", "platform-operations"),
        ("跨境电商合规管理课程-逐课教案文稿（第1-4课）.md", "跨境电商合规管理", "CM-101", "compliance"),
        ("跨境电商海外品牌管理课程-逐课教案文稿（第1-4课）.md", "跨境电商海外品牌管理", "BM-101", "brand-management"),
        ("跨境物流履约与海外仓体系实战课程-逐课教案文稿（第1-5课）.md", "跨境物流履约与海外仓体系实战", "LW-101", "logistics"),
    ]
    for fname, cname, code, folder in lessons:
        build_pdf(SRC / fname, cname, code, OUT / folder / "lesson-plan-sample.pdf")
    build_logistics_cost_xlsx(OUT / "supply-chain" / "logistics-cost.xlsx")
    build_supplier_xlsx(OUT / "supply-chain" / "supplier-evaluation.xlsx")
    build_toolkit_zip(OUT / "toolkit" / "toolkit-samples.zip")
    print("全部 8 个资源文件生成完成。")


if __name__ == "__main__":
    sys.exit(main())
